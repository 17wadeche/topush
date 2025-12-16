from __future__ import annotations
import json
import os
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None
    get_column_letter = None
CENTRAL_TZ = ZoneInfo("America/Chicago")
DEFAULT_TELEMETRY_DIR = r"\\hcwda30449e\Validation-Tool\logs"
DEFAULT_TELEMETRY_FILE = "telemetry.jsonl"
ERROR_COLUMNS: List[Tuple[str, Callable[[Dict[str, Any]], Any]]] = [
    ("Date", lambda e: e.get("date", "")),
    ("Time", lambda e: e.get("time", "")),
    ("user", lambda e: e.get("user", "")),
    ("action", lambda e: e.get("action", "")),
    ("app_version", lambda e: e.get("app_version", "")),
    ("model", lambda e: e.get("model", "")),
    ("release", lambda e: e.get("release_type", "")),
    ("error", lambda e: e.get("error", "")),
]
FEEDBACK_COLUMNS: List[Tuple[str, Callable[[Dict[str, Any]], Any]]] = [
    ("Date", lambda e: e.get("date", "")),
    ("Time", lambda e: e.get("time", "")),
    ("user", lambda e: e.get("user", "")),
    ("app_version", lambda e: e.get("app_version", "")),
    ("model", lambda e: e.get("model", "")),
    ("feedback_text", lambda e: (e.get("payload") or {}).get("feedback_text", "")),
    ("release", lambda e: (e.get("payload") or {}).get("release_type", "")),
]
SUCCESS_COLUMNS: List[Tuple[str, Callable[[Dict[str, Any]], Any]]] = [
    ("Date", lambda e: e.get("date", "")),
    ("Time", lambda e: e.get("time", "")),
    ("user", lambda e: e.get("user", "")),
    ("action", lambda e: e.get("action", "")),
    ("app_version", lambda e: e.get("app_version", "")),
    ("model", lambda e: e.get("model", "")),
    ("release", lambda e: e.get("release_type", "")),
    ("missing", lambda e: e.get("missing_placeholders_count", "")),
    ("draft_q", lambda e: e.get("draft_questions_count", "")),
    ("timings_s", lambda e: e.get("timings_clean", "")),
    ("total_s", lambda e: e.get("timings_total_s", "")), 
]
def ms_to_s(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return round(float(v) / 1000.0, 3)
    except Exception:
        return None
def timings_ms_to_s(timings: Any) -> Dict[str, float]:
    if not isinstance(timings, dict):
        return {}
    out: Dict[str, float] = {}
    for k, v in timings.items():
        s = ms_to_s(v)
        if s is not None:
            out[k.replace("_ms", "_s")] = s
    return out
def format_timings_clean(timings_s: Any) -> str:
    if not isinstance(timings_s, dict) or not timings_s:
        return ""
    items: List[str] = []
    total = 0.0
    for k in sorted(timings_s.keys()):
        try:
            v = float(timings_s[k])
        except Exception:
            continue
        total += v
        items.append(f"{k}={v:.3f}")
    items.append(f"total_s={total:.3f}")
    return "\n".join(items)
def split_ts_utc_to_central(ts: Any) -> Tuple[str, str]:
    if not ts:
        return "", ""
    s = str(ts).strip()
    try:
        s2 = s[:-1] + "+00:00" if s.endswith("Z") else s
        dt = datetime.fromisoformat(s2)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        local = dt.astimezone(CENTRAL_TZ)
        return local.date().isoformat(), local.time().replace(microsecond=0).isoformat()
    except Exception:
        return "", ""
def read_jsonl(path: Path, max_lines: int = 5000) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    lines = lines[-max_lines:]
    events: List[Dict[str, Any]] = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        try:
            events.append(json.loads(line))
        except Exception:
            continue
    events.sort(key=lambda r: r.get("ts_utc", ""), reverse=True)
    return events
def read_jsonl_many(dir_path: Path, pattern: str = "telemetry*.jsonl", max_lines_per_file: int = 5000) -> List[Dict[str, Any]]:
    events: List[Dict[str, Any]] = []
    if not dir_path.exists():
        return events
    for p in sorted(dir_path.glob(pattern)):
        events.extend(read_jsonl(p, max_lines=max_lines_per_file))
    events.sort(key=lambda r: r.get("ts_utc", ""), reverse=True)
    return events
def total_timings_s(timings_s: Any) -> float:
    if not isinstance(timings_s, dict):
        return 0.0
    total = 0.0
    for v in timings_s.values():
        try:
            total += float(v)
        except Exception:
            continue
    return round(total, 3)
def split_events(
    events: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    errors: List[Dict[str, Any]] = []
    feedback: List[Dict[str, Any]] = []
    success_runs: List[Dict[str, Any]] = []
    for e in events:
        et = e.get("event_type")
        ok = bool(e.get("success", False))
        payload = e.get("payload") if isinstance(e.get("payload"), dict) else {}
        d, t = split_ts_utc_to_central(e.get("ts_utc"))
        e["date"] = d
        e["time"] = t
        e["release_type"] = payload.get("release_type")
        e["draft_questions_count"] = payload.get("draft_questions_count")
        e["missing_placeholders_count"] = payload.get("missing_placeholders_count")
        e["examples_count"] = payload.get("examples_count")
        t_s = timings_ms_to_s(payload.get("timings_ms"))
        e["timings_s"] = t_s
        e["timings_total_s"] = total_timings_s(t_s)
        e["timings_clean"] = format_timings_clean(t_s)
        if et == "ui_feedback":
            feedback.append(e)
        elif et == "ui_run":
            if ok:
                success_runs.append(e)
            else:
                errors.append(e)
    return errors, feedback, success_runs
def h(s: Any) -> str:
    s = "" if s is None else str(s)
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )
def render_table(
    title: str,
    rows: List[Dict[str, Any]],
    columns: List[Tuple[str, Callable[[Dict[str, Any]], Any]]],
    *,
    pill_class: str = "pill",
) -> str:
    header_html = "".join(f"<th>{h(col_name)}</th>" for col_name, _ in columns)
    body_parts: List[str] = []
    for r in rows:
        tds: List[str] = []
        for _, fn in columns:
            val = fn(r)
            if isinstance(val, str) and "\n" in val:
                tds.append(f'<td><code class="mono">{h(val)}</code></td>')
            else:
                if isinstance(val, str) and (val.startswith("20") or ":" in val):
                    tds.append(f"<td><code>{h(val)}</code></td>")
                else:
                    tds.append(f"<td>{h(val)}</td>")
        body_parts.append("<tr>" + "".join(tds) + "</tr>")
    return f"""
  <h2>{h(title)} <span class="{pill_class}">{len(rows)}</span></h2>
  <div class="wrap">
    <table>
      <thead><tr>{header_html}</tr></thead>
      <tbody>
        {''.join(body_parts)}
      </tbody>
    </table>
  </div>
"""
def build_html(
    errors: List[Dict[str, Any]],
    feedback: List[Dict[str, Any]],
    success_runs: List[Dict[str, Any]],
    src_path: Path,
) -> str:
    now = datetime.utcnow().isoformat() + "Z"
    downloads = """
    <div class="muted" style="margin: 8px 0 14px;">
      <div><strong>Exports (Excel):</strong>
        <a href="telemetry_errors.xlsx">Errors</a> •
        <a href="telemetry_feedback.xlsx">Feedback</a> •
        <a href="telemetry_success.xlsx">Successful Runs</a>
      </div>
    </div>
    """
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Validation Logs</title>
  <style>
    body {{ font-family: system-ui, -apple-system, sans-serif; padding: 18px; background:#f8fafc; color:#0f172a; }}
    h1 {{ margin: 0 0 6px; }}
    h2 {{ margin-top: 24px; }}
    .muted {{ color:#475569; }}
    .pill {{ display:inline-block; padding:3px 8px; border-radius:999px; background:#eef2ff; border:1px solid rgba(37,99,235,0.18); color:#334155; font-size:12px; }}
    .errpill {{ background:#fee2e2; border-color:#fecaca; color:#991b1b; }}
    table {{
      width:100%;
      border-collapse: collapse;
      background:#fff;
      border:1px solid rgba(15,23,42,0.12);
      border-radius: 10px;
      overflow:hidden;
      table-layout: fixed; /* required for draggable column widths */
    }}
    th, td {{
      text-align:left;
      padding:10px 12px;
      border-bottom:1px solid rgba(15,23,42,0.08);
      vertical-align: top;
      font-size: 13px;
      overflow: hidden;
    }}
    th {{
      background:#f1f5f9;
      font-size: 13px;
      position: relative;
      user-select: none;
    }}
    th .th-inner {{
      display:flex;
      align-items:center;
      gap:8px;
    }}
    th .resizer {{
      position:absolute;
      right:0;
      top:0;
      width:8px;
      height:100%;
      cursor: col-resize;
    }}
    th.resizing {{
      cursor: col-resize;
    }}
    code {{ font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace; font-size: 12px; }}
    .mono {{ white-space: pre; }}
    .wrap {{ max-width: 100%; overflow-x:auto; }}
  </style>
</head>
<body>
  <h1>Validation Logs</h1>
  {downloads}
  {render_table("Errors", errors, ERROR_COLUMNS, pill_class="pill errpill")}
  {render_table("Feedback", feedback, FEEDBACK_COLUMNS)}
  {render_table("Successful Runs", success_runs, SUCCESS_COLUMNS)}
  <script>
  (function () {{
    function makeResizable(table) {{
      const ths = table.querySelectorAll("thead th");
      ths.forEach((th) => {{
        if (!th.querySelector(".th-inner")) {{
          const wrap = document.createElement("div");
          wrap.className = "th-inner";
          while (th.firstChild) wrap.appendChild(th.firstChild);
          th.appendChild(wrap);
        }}
        const resizer = document.createElement("div");
        resizer.className = "resizer";
        th.appendChild(resizer);
        let startX = 0;
        let startW = 0;
        resizer.addEventListener("mousedown", (e) => {{
          e.preventDefault();
          startX = e.clientX;
          startW = th.getBoundingClientRect().width;
          th.classList.add("resizing");
          const onMove = (ev) => {{
            const dx = ev.clientX - startX;
            const w = Math.max(60, startW + dx);
            th.style.width = w + "px";
          }};
          const onUp = () => {{
            th.classList.remove("resizing");
            window.removeEventListener("mousemove", onMove);
            window.removeEventListener("mouseup", onUp);
          }};
          window.addEventListener("mousemove", onMove);
          window.addEventListener("mouseup", onUp);
        }});
      }});
    }}
    document.querySelectorAll("table").forEach(makeResizable);
  }})();
  </script>
</body>
</html>
"""
def _stringify_excel(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, (dict, list, tuple)):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)
    return str(v)
def write_xlsx(path: Path, rows: List[Dict[str, Any]], columns: List[Tuple[str, Callable[[Dict[str, Any]], Any]]], sheet: str) -> None:
    if openpyxl is None or get_column_letter is None:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    headers = [name for name, _ in columns]
    ws.append(headers)
    for r in rows:
        ws.append([_stringify_excel(fn(r)) for _, fn in columns])
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, min(len(rows) + 2, 500)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            s = str(val)
            s0 = s.splitlines()[0] if "\n" in s else s
            if len(s0) > max_len:
                max_len = len(s0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 70)
    wb.save(path)
def main() -> None:
    dir_path = Path(os.environ.get("TELEMETRY_DIR", DEFAULT_TELEMETRY_DIR))
    pattern = os.environ.get("TELEMETRY_GLOB", "telemetry*.jsonl")
    events = read_jsonl_many(dir_path, pattern=pattern, max_lines_per_file=20000)
    errors, feedback, success_runs = split_events(events)
    out_dir = Path.cwd()
    write_xlsx(out_dir / "telemetry_errors.xlsx", errors, ERROR_COLUMNS, sheet="Errors")
    write_xlsx(out_dir / "telemetry_feedback.xlsx", feedback, FEEDBACK_COLUMNS, sheet="Feedback")
    write_xlsx(out_dir / "telemetry_success.xlsx", success_runs, SUCCESS_COLUMNS, sheet="Successful Runs")
    out = out_dir / "telemetry_report.html"
    out.write_text(build_html(errors, feedback, success_runs, dir_path), encoding="utf-8")
    webbrowser.open(out.as_uri())
    print(f"Wrote: {out}")
if __name__ == "__main__":
    main()